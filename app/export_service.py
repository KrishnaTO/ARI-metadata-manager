"""Export the current ontology to an Excel report in the 1_Core_ARI_Diseases.xlsx
format, optionally marking what changed versus a baseline (the source branch).
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMNS = ["ARI ID", "IRI", "Preferred Name", "Synonyms", "Subtypes", "SNOMED Code(s)",
           "Obsolete SNOMED", "OMOP ConceptID", "Concept Code (DXCODE)", "Code Status",
           "Definition", "Definition Source(s)", "Tissue Region", "Evidence Level",
           "Autoimmune Modifier", "Version"]

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name=FONT, size=10)
CHG_FILL = PatternFill("solid", fgColor="FCE4D6")   # amber: changed cell
NEW_FILL = PatternFill("solid", fgColor="E2EFDA")   # green: new disease
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _s(v):
    """Stringify scalar / list / list-of-dicts(name) annotation values."""
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        parts = []
        for x in v:
            if isinstance(x, dict):
                parts.append(str(x.get("name") or x.get("label") or x.get("iri") or ""))
            else:
                parts.append(str(x))
        return "; ".join(p for p in parts if p)
    return str(v)


def _subtypes(d: dict) -> str:
    """Human-readable subtypes cell: ``Name - description`` with ``→ Linked
    disease`` appended when the subtype links to an existing disease. Falls back
    to the raw annotation strings if the parsed form is unavailable."""
    parsed = d.get("clinical_subtypes_parsed")
    if not parsed:
        return _s(d.get("clinical_subtypes"))
    parts = []
    for s in parsed:
        t = s.get("name") or ""
        if s.get("description"):
            t += f" - {s['description']}"
        if s.get("link_name"):
            t += f" → {s['link_name']}"
        if t:
            parts.append(t)
    return "; ".join(parts)


def disease_to_row(d: dict) -> dict:
    snomed = _s(d.get("snomed"))
    omop = _s(d.get("omop"))
    dx = _s(d.get("dxcode"))
    has_code = bool(snomed or omop or dx)
    return {
        "ARI ID": _s(d.get("ari_id")),
        "IRI": _s(d.get("iri")),
        "Preferred Name": _s(d.get("name")),
        "Synonyms": _s(d.get("synonyms")),
        "Subtypes": _subtypes(d),
        "SNOMED Code(s)": snomed,
        "Obsolete SNOMED": snomed if d.get("obsolete") else "",
        "OMOP ConceptID": omop,
        "Concept Code (DXCODE)": dx,
        "Code Status": "" if has_code else "No code",
        "Definition": _s(d.get("definition")),
        "Definition Source(s)": _s(d.get("def_source")),
        "Tissue Region": _s(d.get("category")),
        "Evidence Level": _s(d.get("evidence")),
        "Autoimmune Modifier": _s(d.get("autoimmune")),
        "Version": _s(d.get("version")),
    }


def _rows_from_service(service) -> dict:
    """Map ARI ID (or IRI fallback) -> core-format row dict."""
    out = {}
    for item in service.get_diseases_list():
        detail = service.get_disease_detail(item["iri"])
        row = disease_to_row(detail)
        out[row["ARI ID"] or row["IRI"]] = row
    return out


def build_report(current_service, baseline_service=None) -> bytes:
    """Build the xlsx. If baseline_service is given, mark changed cells and add a
    'Change Status' column (New / Modified / Unchanged)."""
    cur = _rows_from_service(current_service)
    base = _rows_from_service(baseline_service) if baseline_service is not None else None

    headers = (["Change Status"] if base is not None else []) + COLUMNS
    wb = Workbook(); ws = wb.active; ws.title = "Core ARI Diseases"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = BORDER
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 28

    offset = 1 if base is not None else 0
    for key in sorted(cur, key=lambda k: cur[k]["Preferred Name"].lower()):
        row = cur[key]
        status, changed = "", set()
        if base is not None:
            if key not in base:
                status = "New"
            else:
                changed = {c for c in COLUMNS if row[c] != base[key][c]}
                status = "Modified" if changed else "Unchanged"
        values = ([status] if base is not None else []) + [row[c] for c in COLUMNS]
        ws.append(values)
        r = ws.max_row
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=ci); cell.font = CELL_FONT; cell.border = BORDER
            cell.alignment = WRAP if headers[ci - 1] in ("Synonyms", "Subtypes", "Definition", "Definition Source(s)") else TOP
        if base is not None:
            if status == "New":
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=r, column=ci).fill = NEW_FILL
            else:
                for col in changed:
                    ws.cell(row=r, column=COLUMNS.index(col) + 1 + offset).fill = CHG_FILL

    widths = ([14] if base is not None else []) + [14, 50, 30, 32, 32, 20, 16, 16, 18, 12, 60, 26, 16, 14, 16, 9]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
