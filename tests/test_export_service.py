"""Coverage for the Excel export builder."""
import io

import openpyxl

from app import export_service as ex


# --------------------------------------------------------------- pure helpers
def test_stringify_scalar_list_and_dicts():
    assert ex._s(None) == ""
    assert ex._s(["a", "b"]) == "a; b"
    assert ex._s([{"name": "x"}, {"label": "y"}, {"iri": "z"}]) == "x; y; z"
    assert ex._s("plain") == "plain"


def test_disease_to_row_marks_missing_codes():
    row = ex.disease_to_row({"name": "X", "iri": "i", "snomed": [], "omop": [], "dxcode": []})
    assert row["Code Status"] == "No code"
    assert row["Preferred Name"] == "X"

    coded = ex.disease_to_row({"name": "Y", "snomed": ["12345"]})
    assert coded["Code Status"] == ""
    assert coded["SNOMED Code(s)"] == "12345"


def test_subtypes_render_with_links():
    d = {"clinical_subtypes_parsed": [
        {"name": "Juvenile", "description": "early onset", "link_name": "Child disease"},
    ]}
    out = ex._subtypes(d)
    assert "Juvenile" in out and "early onset" in out and "Child disease" in out


# --------------------------------------------------------------- integration
def _load(xlsx_bytes):
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes)).active


def test_build_report_without_baseline(ro_service):
    ws = _load(ex.build_report(ro_service))
    headers = [c.value for c in ws[1]]
    assert "Preferred Name" in headers
    assert "Change Status" not in headers          # no baseline -> no status column
    assert ws.max_row > 1                           # at least one disease row


def test_build_report_with_baseline_flags_changes(make_service):
    baseline = make_service()
    current = make_service()
    iri = current.get_diseases_list()[0]["iri"]
    current.update_disease(iri, {"disease_category": "ZZZ-Export-Test"}, editor="t")

    ws = _load(ex.build_report(current, baseline))
    headers = [c.value for c in ws[1]]
    assert headers[0] == "Change Status"
    statuses = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert "Modified" in statuses


def test_build_report_with_baseline_flags_new_disease(make_service):
    baseline = make_service()
    current = make_service()
    current.create_disease({"label": "Exportable New Disease"}, editor="t")

    ws = _load(ex.build_report(current, baseline))
    statuses = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert "New" in statuses
